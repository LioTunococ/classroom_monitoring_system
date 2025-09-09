from calendar import monthrange
import calendar as _cal
from datetime import date, timedelta
from django.core.cache import cache
import csv
from io import TextIOWrapper

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q, Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.contrib.auth import get_user_model

from .forms import AttendanceFormSet, SchoolYearForm, StudentForm, PeriodForm
from .permissions import has_feature
from .models import AttendanceSessionRecord, Enrollment, SchoolYear, Student, Section, NonSchoolDay, Notification, Period, AttendancePeriodRecord, SectionAccess

# Status codes used across reports and dashboard
STATUS_CODES = ('P', 'A', 'L', 'E')
PRESENT_SET = {'P', 'L', 'E'}  # Treat Late and Excused as present


def _first_friday_of_sy(sy: SchoolYear):
    d = sy.start_date
    while d.weekday() != 4:  # Monday=0 ... Friday=4
        d += timedelta(days=1)
    return d


def _school_days(days, sy: SchoolYear):
    # Weekdays only, excluding declared NonSchoolDay for this school year within the given days
    weekdays = [d for d in days if d.weekday() < 5]
    if not weekdays:
        return weekdays
    start, end = min(weekdays), max(weekdays)
    excluded = set(NonSchoolDay.objects.filter(school_year=sy, date__gte=start, date__lte=end).values_list('date', flat=True))
    return [d for d in weekdays if d not in excluded]


def _compute_sf2_summary(sy: SchoolYear, year: int, month: int, days, enrollments, by_key):
    from calendar import monthrange as _mr
    first_day = date(year, month, 1)
    last_day = date(year, month, _mr(year, month)[1])
    range_end = min(last_day, sy.end_date)

    def sex_filter(code):
        return [e for e in enrollments if e.student.sex == code]

    buckets = {
        'M': sex_filter('M'),
        'F': sex_filter('F'),
        'T': list(enrollments),
    }

    school_days = _school_days(days, sy)
    n_school_days = len(school_days)
    first_friday = _first_friday_of_sy(sy)

    def enrol_first_friday(es):
        return sum(1 for e in es if (e.active and e.date_enrolled <= first_friday))

    def late_enrol(es):
        return sum(1 for e in es if (first_day <= e.date_enrolled <= last_day and e.date_enrolled > first_friday))

    def registered_eom(es):
        return sum(1 for e in es if (e.active and e.date_enrolled <= range_end))

    def total_daily_attendance(es):
        total = 0.0
        for d in school_days:
            day_sum = 0.0
            for e in es:
                am = by_key.get((e.id, d, 'AM'))
                pm = by_key.get((e.id, d, 'PM'))
                amv = 1.0 if (am and am.status in PRESENT_SET) else 0.0
                pmv = 1.0 if (pm and pm.status in PRESENT_SET) else 0.0
                day_sum += (amv + pmv) / 2.0
            total += day_sum
        return total

    def absent5_consecutive(es):
        """
        Counts learners with at least 5 consecutive school-day equivalents of absence.
        - Per day, absence value = 0.5 for each 'A' session (AM/PM), capped at 1.0 per day.
        - Any day with absence value == 0 resets the consecutive streak.
        - A streak reaches 5.0 when cumulative absence across successive days >= 5.0.
        """
        cnt = 0
        for e in es:
            streak = 0.0
            flagged = False
            for d in school_days:
                am = by_key.get((e.id, d, 'AM'))
                pm = by_key.get((e.id, d, 'PM'))
                day_abs = 0.0
                if am and am.status == 'A':
                    day_abs += 0.5
                if pm and pm.status == 'A':
                    day_abs += 0.5
                if day_abs > 1.0:
                    day_abs = 1.0
                if day_abs == 0.0:
                    streak = 0.0
                else:
                    streak += day_abs
                    if streak >= 5.0:
                        flagged = True
                        break
            if flagged:
                cnt += 1
        return cnt

    out = {
        'school_days': n_school_days,
        'first_friday': first_friday,
        'by': {},
    }

    for key, es in buckets.items():
        e1 = enrol_first_friday(es)
        late = late_enrol(es)
        reg = registered_eom(es)
        tda = total_daily_attendance(es)
        ada = (tda / n_school_days) if n_school_days else 0.0
        pct_enrol = (reg / e1 * 100.0) if e1 else 0.0
        pct_att = (ada / reg * 100.0) if reg else 0.0
        ab5 = absent5_consecutive(es)
        out['by'][key] = {
            'enrol_first_friday': e1,
            'late_enrol': late,
            'registered_eom': reg,
            'pct_enrol_eom': round(pct_enrol, 2),
            'ada': round(ada, 2),
            'pct_attendance': round(pct_att, 2),
            'absent5': ab5,
        }
    return out


def _get_active_school_year():
    return SchoolYear.objects.filter(is_active=True).order_by('-start_date').first()


def _user_sections_for_sy(user, sy):
    if user.is_staff or user.is_superuser:
        return Section.objects.filter(school_year=sy)
    return Section.objects.filter(school_year=sy, adviser=user)


# Server-side SMS helpers removed (using phone-based SMS only)


@login_required
def dashboard(request):
    sy = _get_active_school_year()
    date_param = request.GET.get('date')
    try:
        if date_param:
            y, m, d = [int(x) for x in date_param.split('-')]
            view_date = date(y, m, d)
        else:
            view_date = date.today()
    except Exception:
        view_date = date.today()
    upcoming_window = view_date + timedelta(days=14)

    upcoming_birthdays = []
    summary = {}
    top_absent = []
    top_late = []
    if sy:
        # Gather birthdays falling within the next 14 days using a days-until calculation
        enrollments_qs = Enrollment.objects.filter(school_year=sy, active=True).select_related('student')
        # Scope by adviser or officer sections if not staff
        is_staffish = (request.user.is_staff or request.user.is_superuser)
        if not is_staffish:
            officer_section_ids = list(SectionAccess.objects.filter(user=request.user, section__school_year=sy).values_list('section_id', flat=True))
            enrollments_qs = enrollments_qs.filter(Q(section__adviser=request.user) | Q(section_id__in=officer_section_ids))
        for enr in enrollments_qs:
            b = enr.student.birthdate
            if not b:
                continue
            # Compute the next birthday occurrence from the view date
            next_bday = b.replace(year=view_date.year)
            if next_bday < view_date:
                next_bday = next_bday.replace(year=view_date.year + 1)
            days_until = (next_bday - view_date).days
            if 0 <= days_until <= 14:
                upcoming_birthdays.append((enr.student, next_bday))

        upcoming_birthdays.sort(key=lambda x: x[1])

        # Summary cards
        total_enrolled = enrollments_qs.count()
        sections_qs = Section.objects.filter(school_year=sy)
        if not is_staffish:
            sections_qs = sections_qs.filter(Q(adviser=request.user) | Q(id__in=officer_section_ids))
        section_count = sections_qs.count()

        # Attendance progress for selected day (session-based)
        recs_today = AttendanceSessionRecord.objects.filter(
            enrollment__in=enrollments_qs, date=view_date
        ).select_related('enrollment__student')
        recorded_sessions = recs_today.count()
        total_sessions = total_enrolled * 2
        remaining_sessions = max(0, total_sessions - recorded_sessions)
        # Status breakdown across sessions
        counts = {k: 0 for k in STATUS_CODES}
        am_counts = {k: 0 for k in STATUS_CODES}
        pm_counts = {k: 0 for k in STATUS_CODES}
        am_lists = {k: [] for k in ('A','L','E')}
        pm_lists = {k: [] for k in ('A','L','E')}
        existing = set()
        for r in recs_today:
            counts[r.status] = counts.get(r.status, 0) + 1
            existing.add((r.enrollment_id, r.session))
            if r.session == 'AM':
                am_counts[r.status] = am_counts.get(r.status, 0) + 1
                if r.status in am_lists:
                    s = r.enrollment.student
                    am_lists[r.status].append({
                        'name': f"{s.last_name}, {s.first_name}",
                        'phone': s.guardian_phone or '',
                    })
            elif r.session == 'PM':
                pm_counts[r.status] = pm_counts.get(r.status, 0) + 1
                if r.status in pm_lists:
                    s = r.enrollment.student
                    pm_lists[r.status].append({
                        'name': f"{s.last_name}, {s.first_name}",
                        'phone': s.guardian_phone or '',
                    })

        # Missing records by session
        missing_lists = {'AM': [], 'PM': []}
        for e in enrollments_qs.select_related('student'):
            name = f"{e.student.last_name}, {e.student.first_name}"
            if (e.id, 'AM') not in existing:
                missing_lists['AM'].append(name)
            if (e.id, 'PM') not in existing:
                missing_lists['PM'].append(name)

        summary = {
            'total_enrolled': total_enrolled,
            'section_count': section_count,
            'recorded_sessions': recorded_sessions,
            'total_sessions': total_sessions,
            'remaining_sessions': remaining_sessions,
            'status_counts': counts,
            'am_counts': am_counts,
            'pm_counts': pm_counts,
            'progress_pct': int((recorded_sessions / total_sessions) * 100) if total_sessions else 0,
            'date': view_date,
            'am_lists': am_lists,
            'pm_lists': pm_lists,
            'missing_lists': missing_lists,
            'is_complete': (recorded_sessions == total_sessions and total_sessions > 0),
        }

        # Top absences and lates for the current month (bounded by School Year)
        month_start = date(view_date.year, view_date.month, 1)
        ms = max(sy.start_date, month_start)
        last_day = monthrange(view_date.year, view_date.month)[1]
        month_end = date(view_date.year, view_date.month, last_day)
        me = min(sy.end_date, month_end)
        agg = (
            AttendanceSessionRecord.objects.filter(
                enrollment__in=enrollments_qs,
                date__gte=ms,
                date__lte=me,
            )
            .values(
                'enrollment__student__id',
                'enrollment__student__last_name',
                'enrollment__student__first_name',
            )
            .annotate(
                abs_sess=Count('id', filter=Q(status='A')),
                late_sess=Count('id', filter=Q(status='L')),
            )
        )

        for row in agg:
            name = f"{row['enrollment__student__last_name']}, {row['enrollment__student__first_name']}"
            sid = row['enrollment__student__id']
            a = int(row.get('abs_sess') or 0)
            l = int(row.get('late_sess') or 0)
            if a > 0:
                top_absent.append({
                    'student_id': sid,
                    'student_name': name,
                    'abs_sessions': a,
                    'days_absent_equiv': round(a / 2.0, 1),
                })
            if l > 0:
                top_late.append({
                    'student_id': sid,
                    'student_name': name,
                    'late_sessions': l,
                })

        # Sort and keep top 5 each
        top_absent.sort(key=lambda x: (x['abs_sessions'], x['student_name']), reverse=True)
        top_late.sort(key=lambda x: (x['late_sessions'], x['student_name']), reverse=True)
        top_absent = top_absent[:5]
        top_late = top_late[:5]
        top_period_label = ms.strftime('%b %Y')
        # Compute month navigation enablement
        cur_month_start = date(view_date.year, view_date.month, 1)
        prev_month = (cur_month_start - timedelta(days=1)).replace(day=1)
        next_month = (date(view_date.year + (1 if view_date.month == 12 else 0), (1 if view_date.month == 12 else view_date.month + 1), 1))
        can_prev_month = (prev_month >= sy.start_date)
        can_next_month = (next_month <= sy.end_date)

    context = {
        'active_sy': sy,
        'today': date.today(),
        'view_date': view_date,
        'upcoming_birthdays': upcoming_birthdays,
        'summary': summary,
        'sections': list(sections_qs) if sy else [],
        'top_absent': top_absent,
        'top_late': top_late,
        'top_period_label': top_period_label if sy else None,
        'can_prev_month': can_prev_month if sy else False,
        'can_next_month': can_next_month if sy else False,
    }
    return render(request, 'attendance/dashboard.html', context)


@login_required
def student_list(request):
    if not has_feature(request.user, 'manage_students'):
        messages.warning(request, 'You are not allowed to view Students.')
        return redirect('attendance:dashboard')
    show_archived = request.GET.get('archived') == '1'
    qs = Student.objects.all()
    if not show_archived:
        qs = qs.filter(is_active=True)
    students = qs
    try:
        active_sy = SchoolYear.objects.filter(is_active=True).first()
    except Exception:
        active_sy = None
    return render(request, 'attendance/students_list.html', {
        'students': students,
        'show_archived': show_archived,
        'active_sy': active_sy,
    })


@login_required
def student_create(request):
    if not has_feature(request.user, 'manage_students'):
        messages.warning(request, 'You are not allowed to add Students.')
        return redirect('attendance:dashboard')
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            student = form.save()
            messages.success(request, f'Student {student} created.')
            return redirect('attendance:student_list')
    else:
        form = StudentForm()
    return render(request, 'attendance/student_form.html', {'form': form})


@login_required
def student_edit(request, pk: int):
    if not has_feature(request.user, 'manage_students'):
        messages.warning(request, 'You are not allowed to edit Students.')
        return redirect('attendance:dashboard')
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            student = form.save()
            messages.success(request, f'Student {student} updated.')
            return redirect('attendance:student_list')
    else:
        form = StudentForm(instance=student)
    return render(request, 'attendance/student_form.html', {'form': form})


@login_required

@login_required
def student_delete(request, pk: int):
    if not has_feature(request.user, 'manage_students'):
        messages.warning(request, 'You are not allowed to delete Students.')
        return redirect('attendance:dashboard')
    student = get_object_or_404(Student, pk=pk)
    from .models import Enrollment, AttendanceSessionRecord
    enroll_count = Enrollment.objects.filter(student=student).count()
    attn_count = AttendanceSessionRecord.objects.filter(enrollment__student=student).count()
    if request.method == 'POST':
        # Allow hard delete only if superuser OR no linked data
        if request.user.is_superuser or (enroll_count == 0 and attn_count == 0):
            name = str(student)
            student.delete()
            messages.success(request, f'Student {name} deleted.')
        else:
            messages.error(request, 'Cannot hard delete a student with enrollments or attendance. Archive instead.')
        return redirect('attendance:student_list')
    return render(request, 'attendance/student_confirm_delete.html', {
        'student': student,
        'enroll_count': enroll_count,
        'attn_count': attn_count,
    })


@login_required
def student_archive(request, pk: int):
    if not has_feature(request.user, 'manage_students'):
        messages.warning(request, 'You are not allowed to archive Students.')
        return redirect('attendance:dashboard')
    student = get_object_or_404(Student, pk=pk)
    student.is_active = False
    student.save(update_fields=["is_active"])
    messages.success(request, f'Student {student} archived.')
    return redirect('attendance:student_list')


@login_required
def student_restore(request, pk: int):
    if not has_feature(request.user, 'manage_students'):
        messages.warning(request, 'You are not allowed to restore Students.')
        return redirect('attendance:dashboard')
    student = get_object_or_404(Student, pk=pk)
    student.is_active = True
    student.save(update_fields=["is_active"])
    messages.success(request, f'Student {student} restored.')
    return redirect('attendance:student_list')


@login_required
def student_history(request, pk: int):
    if not has_feature(request.user, 'view_student_history'):
        messages.warning(request, 'You are not allowed to view student history.')
        return redirect('attendance:dashboard')
    student = get_object_or_404(Student, pk=pk)

    sy_param = request.GET.get('schoolyear_id')
    try:
        sy = get_object_or_404(SchoolYear, pk=int(sy_param)) if sy_param else _get_active_school_year()
    except Exception:
        sy = _get_active_school_year()
    if not sy:
        messages.error(request, 'No active school year found.')
        return redirect('attendance:dashboard')
    today = date.today()
    try:
        year = int(request.GET.get('year') or today.year)
        month = int(request.GET.get('month') or today.month)
    except Exception:
        year, month = today.year, today.month

    enrollment = Enrollment.objects.filter(student=student, school_year=sy, active=True).select_related('section').first()
    if not enrollment:
        messages.info(request, f'{student} is not enrolled in {sy.name}.')
        return render(request, 'attendance/student_history.html', {
            'student': student, 'schoolyear': sy, 'year': year, 'month': month, 'days': [], 'entries': [], 'counts': {},
        })

    first_day = date(year, month, 1)
    last_day = date(year, month, monthrange(year, month)[1])
    range_start = max(first_day, sy.start_date)
    range_end = min(last_day, sy.end_date)
    if range_start > range_end:
        days = []
    else:
        days = [range_start + timedelta(n) for n in range((range_end - range_start).days + 1)]

    recs = AttendanceSessionRecord.objects.filter(enrollment=enrollment, date__gte=range_start, date__lte=range_end)
    by_key = {(r.date, r.session): r for r in recs}
    entries = []
    counts = {'P': 0.0, 'A': 0.0, 'L': 0.0, 'E': 0.0}
    for d in days:
        am = by_key.get((d, 'AM'))
        pm = by_key.get((d, 'PM'))
        ams = am.status if am else ''
        pms = pm.status if pm else ''
        for s in (ams, pms):
            if s:
                counts[s] = counts.get(s, 0.0) + 0.5
                if s in PRESENT_SET and s != 'P':
                    counts['P'] += 0.5
        severity = 'ok'
        if ams == 'A' or pms == 'A':
            severity = 'abs'
        elif ams == 'L' or pms == 'L':
            severity = 'late'
        elif ams == 'E' or pms == 'E':
            severity = 'exc'
        entries.append({
            'day': d,
            'am': ams,
            'pm': pms,
            'remarks': ', '.join(filter(None, [getattr(am, 'remarks', ''), getattr(pm, 'remarks', '')])),
            'sev': severity,
        })

    non_school_days = set(NonSchoolDay.objects.filter(school_year=sy, date__gte=range_start, date__lte=range_end).values_list('date', flat=True))
    # Build calendar grid (Mon-Sun)
    month_first = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    month_last = date(year, month, last_day)
    # Monday=0
    grid_start = month_first - timedelta(days=month_first.weekday())
    grid_end = month_last + timedelta(days=(6 - month_last.weekday()))
    grid_days = []
    d = grid_start
    while d <= grid_end:
        grid_days.append(d)
        d += timedelta(days=1)
    weeks = [grid_days[i:i+7] for i in range(0, len(grid_days), 7)]
    entry_by_date = {e['day']: e for e in entries}

    return render(request, 'attendance/student_history.html', {
        'student': student,
        'schoolyear': sy,
        'enrollment': enrollment,
        'year': year,
        'month': month,
        'days': days,
        'entries': entries,
        'counts': counts,
        'nsd_dates': non_school_days,
        'weeks': weeks,
        'entry_by_date': entry_by_date,
        'grid_start': grid_start,
        'grid_end': grid_end,
    })

@login_required
def schoolyear_list(request):
    if not has_feature(request.user, 'manage_schoolyears'):
        messages.warning(request, 'You are not allowed to view School Years.')
        return redirect('attendance:dashboard')
    sys = SchoolYear.objects.all()
    return render(request, 'attendance/schoolyear_list.html', {'schoolyears': sys})


@login_required
def schoolyear_create(request):
    if not has_feature(request.user, 'manage_schoolyears'):
        messages.warning(request, 'You are not allowed to create School Years.')
        return redirect('attendance:dashboard')
    if request.method == 'POST':
        form = SchoolYearForm(request.POST)
        if form.is_valid():
            sy = form.save()
            if sy.is_active:
                SchoolYear.objects.exclude(pk=sy.pk).update(is_active=False)
            messages.success(request, f'School Year {sy.name} created.')
            return redirect('attendance:schoolyear_list')
    else:
        form = SchoolYearForm()
    return render(request, 'attendance/schoolyear_form.html', {'form': form})


@login_required
def schoolyear_edit(request, pk: int):
    if not has_feature(request.user, 'manage_schoolyears'):
        messages.warning(request, 'You are not allowed to edit School Years.')
        return redirect('attendance:dashboard')
    sy = get_object_or_404(SchoolYear, pk=pk)
    if request.method == 'POST':
        form = SchoolYearForm(request.POST, instance=sy)
        if form.is_valid():
            sy = form.save()
            if sy.is_active:
                SchoolYear.objects.exclude(pk=sy.pk).update(is_active=False)
            messages.success(request, f'School Year {sy.name} updated.')
            return redirect('attendance:schoolyear_list')
    else:
        form = SchoolYearForm(instance=sy)
    return render(request, 'attendance/schoolyear_form.html', {'form': form})


@login_required
def enroll_students(request, schoolyear_id: int):
    if not has_feature(request.user, 'enroll_students'):
        messages.warning(request, 'You are not allowed to manage enrollment.')
        return redirect('attendance:dashboard')
    sy = get_object_or_404(SchoolYear, pk=schoolyear_id)
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        created = 0
        for sid in student_ids:
            student = get_object_or_404(Student, pk=sid)
            _, was_created = Enrollment.objects.get_or_create(student=student, school_year=sy)
            if was_created:
                created += 1
        messages.success(request, f'Enrolled {created} new student(s) to {sy.name}.')
        # Go straight to taking attendance for convenience
        return redirect('attendance:take_attendance', schoolyear_id=sy.id)

    enrolled_ids = set(Enrollment.objects.filter(school_year=sy).values_list('student_id', flat=True))
    students = Student.objects.filter(is_active=True)
    return render(request, 'attendance/enroll_students.html', {
        'schoolyear': sy,
        'students': students,
        'enrolled_ids': enrolled_ids,
    })


@login_required
def take_attendance(request, schoolyear_id: int):
    if not has_feature(request.user, 'take_attendance'):
        messages.warning(request, 'You are not allowed to take attendance.')
        return redirect('attendance:dashboard')
    sy = get_object_or_404(SchoolYear, pk=schoolyear_id)
    target_date_str = request.GET.get('date') or request.POST.get('date')
    if target_date_str:
        y, m, d = [int(x) for x in target_date_str.split('-')]
        target_date = date(y, m, d)
    else:
        target_date = date.today()

    enroll_qs = Enrollment.objects.filter(school_year=sy, active=True).select_related('student', 'section')
    # Restrict to allowed sections for non-staff: adviser or officer access
    if not (request.user.is_staff or request.user.is_superuser):
        officer_section_ids = list(
            SectionAccess.objects.filter(user=request.user, section__school_year=sy).values_list('section_id', flat=True)
        )
        enroll_qs = enroll_qs.filter(
            Q(section__adviser=request.user) | Q(section_id__in=officer_section_ids)
        )
        if not enroll_qs.exists():
            messages.error(request, 'No assigned section or no enrolled students for you in this school year.')
            return redirect('attendance:schoolyear_list')
    enrollments = list(enroll_qs)
    periods_all = list(Period.objects.filter(school_year=sy, is_active=True).order_by('order', 'id'))
    has_periods = len(periods_all) > 0
    sess_qs = AttendanceSessionRecord.objects.filter(enrollment__in=enrollments, date=target_date)
    existing = {}
    for rec in sess_qs:
        existing[(rec.enrollment_id, rec.session)] = rec

    initial = []
    enrollments_period = []
    if not has_periods:
        for e in enrollments:
            student = e.student
            am_rec = existing.get((e.id, 'AM'))
            pm_rec = existing.get((e.id, 'PM'))
            status_am = am_rec.status if am_rec else 'P'
            status_pm = pm_rec.status if pm_rec else 'P'
            remarks = (am_rec.remarks if am_rec else (pm_rec.remarks if pm_rec else ''))
            initial.append({
                'enrollment_id': e.id,
                'student_id': student.id,
                'student_name': f"{student.last_name}, {student.first_name}",
                'guardian_phone': getattr(student, 'guardian_phone', ''),
                'guardian_name': getattr(student, 'guardian_name', ''),
                'status_am': status_am,
                'status_pm': status_pm,
                'remarks': remarks,
            })
    else:
        # Build per-period rendering data with existing records prefilled
        recs = AttendancePeriodRecord.objects.filter(
            enrollment__in=enrollments, date=target_date
        ).select_related('period')
        by_key = {(r.enrollment_id, r.period_id): r for r in recs}
        for e in enrollments:
            student = e.student
            item = {
                'enrollment_id': e.id,
                'student_id': student.id,
                'student_name': f"{student.last_name}, {student.first_name}",
                'guardian_phone': getattr(student, 'guardian_phone', ''),
                'statuses': {},
                'time_in': {},
            }
            for p in periods_all:
                r = by_key.get((e.id, p.id))
                item['statuses'][p.id] = (r.status if r else 'P')
                item['time_in'][p.id] = (r.time_in.strftime('%H:%M') if (r and r.time_in) else '')
            enrollments_period.append(item)

    if request.method == 'POST' and not has_periods:
        formset = AttendanceFormSet(request.POST, initial=initial, prefix='att')
        if formset.is_valid():
            with transaction.atomic():
                for form in formset:
                    eid = form.cleaned_data['enrollment_id']
                    status_am = form.cleaned_data['status_am']
                    status_pm = form.cleaned_data['status_pm']
                    remarks = form.cleaned_data.get('remarks', '')
                    # Keep previous to detect changes
                    prev_am = existing.get((eid, 'AM'))
                    prev_pm = existing.get((eid, 'PM'))
                    am_obj, _ = AttendanceSessionRecord.objects.update_or_create(
                        enrollment_id=eid, date=target_date, session='AM',
                        defaults={'status': status_am, 'remarks': remarks}
                    )
                    pm_obj, _ = AttendanceSessionRecord.objects.update_or_create(
                        enrollment_id=eid, date=target_date, session='PM',
                        defaults={'status': status_pm, 'remarks': remarks}
                    )
                    # Create in-app notifications for non-Present statuses when new or changed
                    try:
                        e = next((x for x in enrollments if x.id == eid), None)
                        student = e.student if e else None
                        if student:
                            def status_word(code):
                                return {'P':'Present','A':'Absent','L':'Late','E':'Excused'}.get(code, code)
                            base_url = f"{reverse('attendance:take_attendance', args=[schoolyear_id])}?date={target_date}"
                            if status_am in {'A','L','E'} and ((not prev_am) or prev_am.status != status_am):
                                Notification.objects.create(
                                    user=request.user,
                                    message=f"{student.last_name}, {student.first_name} is {status_word(status_am)} (AM) on {target_date}",
                                    url=base_url,
                                )
                            if status_pm in {'A','L','E'} and ((not prev_pm) or prev_pm.status != status_pm):
                                Notification.objects.create(
                                    user=request.user,
                                    message=f"{student.last_name}, {student.first_name} is {status_word(status_pm)} (PM) on {target_date}",
                                    url=base_url,
                                )
                    except Exception:
                        pass
            # Invalidate cached monthly summaries for this SY/month
            try:
                year = target_date.year
                month = target_date.month
                # 'all' scope (staff views)
                cache.delete(f"sf2:{sy.id}:{year}:{month}:all")
                # Adviser scopes for impacted sections
                adviser_ids = set(e.section.adviser_id for e in enrollments if e.section_id)
                # Include current user in case they are an adviser
                adviser_ids.add(getattr(request.user, 'id', None))
                for aid in adviser_ids:
                    if aid:
                        cache.delete(f"sf2:{sy.id}:{year}:{month}:user:{aid}")
            except Exception:
                # Cache is best-effort; ignore failures
                pass
            messages.success(request, f"Attendance successfully saved for {target_date.strftime('%B %d, %Y') }.")
            # Redirect to dashboard and keep the selected date context
            return redirect(f"{reverse('attendance:dashboard')}?date={target_date}")
    elif request.method == 'POST' and has_periods:
        from django.db import transaction
        with transaction.atomic():
            for e in enrollments:
                for p in periods_all:
                    status = request.POST.get(f"p_{e.id}_{p.id}_status") or 'P'
                    time_in = request.POST.get(f"ti_{e.id}_{p.id}") or None
                    AttendancePeriodRecord.objects.update_or_create(
                        enrollment=e, date=target_date, period=p,
                        defaults={'status': status, 'time_in': time_in or None}
                    )
                def agg(half):
                    qs = AttendancePeriodRecord.objects.filter(enrollment=e, date=target_date, period__half=half)
                    n = qs.count()
                    if n == 0:
                        return 'P'
                    absent = sum(1 for r in qs if r.status == 'A')
                    late_any = any(r.status == 'L' for r in qs)
                    exc_any = any(r.status == 'E' for r in qs)
                    if absent * 2 >= n:
                        return 'A'
                    if late_any:
                        return 'L'
                    if exc_any:
                        return 'E'
                    return 'P'
                am_status = agg('AM')
                pm_status = agg('PM')
                AttendanceSessionRecord.objects.update_or_create(
                    enrollment=e, date=target_date, session='AM', defaults={'status': am_status}
                )
                AttendanceSessionRecord.objects.update_or_create(
                    enrollment=e, date=target_date, session='PM', defaults={'status': pm_status}
                )
        messages.success(request, f"Attendance successfully saved for {target_date.strftime('%B %d, %Y') }.")
        nav = request.POST.get('nav')
        if nav == 'prev':
            next_date = target_date - timedelta(days=1)
            return redirect(f"{reverse('attendance:take_attendance', args=[sy.id])}?date={next_date}")
        if nav == 'next':
            next_date = target_date + timedelta(days=1)
            return redirect(f"{reverse('attendance:take_attendance', args=[sy.id])}?date={next_date}")
        return redirect(f"{reverse('attendance:dashboard')}?date={target_date}")
    else:
        formset = AttendanceFormSet(initial=initial, prefix='att')

    context = {
        'schoolyear': sy,
        'target_date': target_date,
        'formset': formset,
    }
    if has_periods:
        context.update({
            'periods': periods_all,
            'periods_am': [p for p in periods_all if p.half == 'AM'],
            'periods_pm': [p for p in periods_all if p.half == 'PM'],
            'enrollments_period': enrollments_period,
        })
    return render(request, 'attendance/attendance_form.html', context)

@login_required
def manage_periods(request, schoolyear_id: int):
    if not has_feature(request.user, 'manage_periods'):
        messages.warning(request, 'You are not allowed to manage periods.')
        return redirect('attendance:dashboard')
    sy = get_object_or_404(SchoolYear, pk=schoolyear_id)
    if request.method == 'POST':
        form = PeriodForm(request.POST)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.school_year = sy
            obj.save()
            messages.success(request, f'Period {obj.name} added.')
            return redirect('attendance:manage_periods', schoolyear_id=sy.id)
    else:
        form = PeriodForm()
    periods = Period.objects.filter(school_year=sy).order_by('order', 'id')
    return render(request, 'attendance/periods.html', {'schoolyear': sy, 'periods': periods, 'form': form})


@login_required
def edit_period(request, schoolyear_id: int, pk: int):
    if not has_feature(request.user, 'manage_periods'):
        messages.warning(request, 'You are not allowed to edit periods.')
        return redirect('attendance:dashboard')
    sy = get_object_or_404(SchoolYear, pk=schoolyear_id)
    period = get_object_or_404(Period, pk=pk, school_year=sy)
    if request.method == 'POST':
        form = PeriodForm(request.POST, instance=period)
        if form.is_valid():
            form.save()
            messages.success(request, f'Period {period.name} updated.')
            return redirect('attendance:manage_periods', schoolyear_id=sy.id)
    else:
        form = PeriodForm(instance=period)
    return render(request, 'attendance/period_edit.html', {'schoolyear': sy, 'form': form, 'period': period})


@login_required
def notifications(request):
    qs = Notification.objects.filter(user=request.user).order_by('-created')
    items = list(qs[:100])
    return render(request, 'attendance/notifications.html', {
        'notifications': items,
        'unread_count': qs.filter(is_read=False).count(),
    })


@login_required
def notifications_mark_all_read(request):
    if request.method == 'POST':
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        messages.success(request, 'All notifications marked as read.')
    return redirect('attendance:notifications')


@login_required
def access_users(request):
    # Only admins/advisers (manage_schoolyears) can manage access; superuser ok
    if not (has_feature(request.user, 'manage_schoolyears') or request.user.is_superuser):
        messages.warning(request, 'You are not allowed to manage access.')
        return redirect('attendance:dashboard')
    User = get_user_model()
    q = (request.GET.get('q') or '').strip()
    users = list(User.objects.all().order_by('username'))
    if q:
        ql = q.lower()
        users = [u for u in users if (ql in (u.username or '').lower() or ql in (u.first_name or '').lower() or ql in (u.last_name or '').lower())]
    # Build counts
    from .models import FeatureAccess
    sec_counts = {row['user']: row['c'] for row in SectionAccess.objects.values('user').annotate(c=Count('id'))}
    feat_allow = {}
    feat_deny = {}
    for row in FeatureAccess.objects.values('user', 'allow').annotate(c=Count('id')):
        if row['allow']:
            feat_allow[row['user']] = row['c']
        else:
            feat_deny[row['user']] = row['c']
    items = [{
        'user': u,
        'sec_count': sec_counts.get(u.id, 0),
        'feat_allow': feat_allow.get(u.id, 0),
        'feat_deny': feat_deny.get(u.id, 0),
    } for u in users]
    return render(request, 'attendance/access_users.html', {'items': items, 'q': q})


@login_required
def access_edit(request, user_id: int):
    if not (has_feature(request.user, 'manage_schoolyears') or request.user.is_superuser):
        messages.warning(request, 'You are not allowed to manage access.')
        return redirect('attendance:dashboard')
    User = get_user_model()
    u = get_object_or_404(User, pk=user_id)
    # Sections by SY, features from permissions
    sections = Section.objects.select_related('school_year').order_by('school_year__start_date', 'name')
    from .permissions import FEATURES
    feat_list = sorted(list(FEATURES))

    current_sections = set(SectionAccess.objects.filter(user=u).values_list('section_id', flat=True))
    from .models import FeatureAccess
    current_allow = set(FeatureAccess.objects.filter(user=u, allow=True).values_list('feature', flat=True))
    current_deny = set(FeatureAccess.objects.filter(user=u, allow=False).values_list('feature', flat=True))

    if request.method == 'POST':
        sel_sections = set(int(x) for x in request.POST.getlist('section_ids'))
        allow_feats = set(request.POST.getlist('feat_allow'))
        # Unchecked means deny: compute deny as all - allow
        deny_feats = set(feat_list) - allow_feats
        # Update sections
        to_add = sel_sections - current_sections
        to_del = current_sections - sel_sections
        for sid in to_add:
            try:
                SectionAccess.objects.create(user=u, section_id=sid)
            except Exception:
                pass
        if to_del:
            SectionAccess.objects.filter(user=u, section_id__in=list(to_del)).delete()
        # Update features overrides
        from .models import FeatureAccess
        target = set(feat_list)
        # Delete any override not present anymore (cleanup)
        FeatureAccess.objects.filter(user=u).exclude(feature__in=list(target)).delete()
        # Apply deny first, then allow overrides so allow wins
        for feat in deny_feats:
            FeatureAccess.objects.update_or_create(user=u, feature=feat, defaults={'allow': False})
        for feat in allow_feats:
            FeatureAccess.objects.update_or_create(user=u, feature=feat, defaults={'allow': True})
        messages.success(request, 'Access updated.')
        return redirect('attendance:access_edit', user_id=u.id)

    return render(request, 'attendance/access_edit.html', {
        'the_user': u,
        'sections': sections,
        'current_sections': current_sections,
        'feat_list': feat_list,
        'current_allow': current_allow,
        'current_deny': current_deny,
    })

@login_required
def bulk_assign_section(request, schoolyear_id: int):
    if not has_feature(request.user, 'assign_section'):
        messages.warning(request, 'You are not allowed to assign sections.')
        return redirect('attendance:dashboard')
    sy = get_object_or_404(SchoolYear, pk=schoolyear_id)
    # Sections available to this user in this SY
    if request.user.is_staff or request.user.is_superuser:
        sections = Section.objects.filter(school_year=sy)
        enroll_qs = Enrollment.objects.filter(school_year=sy, active=True).select_related('student', 'section')
    else:
        sections = Section.objects.filter(school_year=sy, adviser=request.user)
        # Non-staff can only modify enrollments with no section or their own section
        enroll_qs = Enrollment.objects.filter(school_year=sy, active=True).filter(Q(section__isnull=True) | Q(section__adviser=request.user)).select_related('student', 'section')

    if request.method == 'POST':
        section_id = request.POST.get('section_id')
        ids = request.POST.getlist('enrollment_ids')
        if not section_id or not ids:
            messages.error(request, 'Please choose a section and at least one student.')
            return redirect(request.path)
        section = get_object_or_404(Section, pk=section_id, school_year=sy)
        # Ensure permission for non-staff
        if not (request.user.is_staff or request.user.is_superuser) and section.adviser_id != request.user.id:
            messages.error(request, 'You cannot assign to a section you do not advise.')
            return redirect(request.path)
        updated = Enrollment.objects.filter(pk__in=ids, school_year=sy, active=True)
        if not (request.user.is_staff or request.user.is_superuser):
            updated = updated.filter(Q(section__isnull=True) | Q(section__adviser=request.user))
        count = updated.update(section=section)
        messages.success(request, f'Assigned section "{section.name}" to {count} student(s).')
        return redirect('attendance:take_attendance', schoolyear_id=sy.id)

    enrollments = list(enroll_qs)
    return render(request, 'attendance/bulk_assign_section.html', {
        'schoolyear': sy,
        'sections': sections,
        'enrollments': enrollments,
    })


@login_required
def report_form(request):
    if not has_feature(request.user, 'view_reports'):
        messages.warning(request, 'You are not allowed to view reports.')
        return redirect('attendance:dashboard')
    sys = SchoolYear.objects.all()
    today = date.today()

    # Resolve selected parameters from query or sensible defaults
    sel_sy_id = request.GET.get('schoolyear_id')
    try:
        sel_sy_id = int(sel_sy_id) if sel_sy_id is not None else None
    except (TypeError, ValueError):
        sel_sy_id = None

    # Prefer active school year, then the first available
    sel_sy = None
    if sel_sy_id:
        sel_sy = get_object_or_404(SchoolYear, pk=sel_sy_id)
    else:
        sel_sy = _get_active_school_year() or sys.first()

    try:
        sel_year = int(request.GET.get('year')) if request.GET.get('year') else today.year
    except (TypeError, ValueError):
        sel_year = today.year
    try:
        sel_month = int(request.GET.get('month')) if request.GET.get('month') else today.month
    except (TypeError, ValueError):
        sel_month = today.month

    # Options for dropdowns
    if sel_sy:
        year_start = sel_sy.start_date.year
        year_end = sel_sy.end_date.year
        year_options = list(range(year_start, year_end + 1))
    else:
        year_options = [today.year - 1, today.year, today.year + 1]

    month_options = [(i, _cal.month_name[i]) for i in range(1, 13)]
    # Optional section filter for admins (staff/superuser)
    sel_section_id = request.GET.get('section_id')
    try:
        sel_section_id = int(sel_section_id) if sel_section_id not in (None, '', 'all') else None
    except (TypeError, ValueError):
        sel_section_id = None

    # Build preview data (days, rows) with the same logic as report_preview
    days = []
    rows = []
    non_school_days = []
    summary = None
    if sel_sy:
        try:
            first_day = date(sel_year, sel_month, 1)
            from calendar import monthrange as _mr
            last_day = date(sel_year, sel_month, _mr(sel_year, sel_month)[1])
            range_start = max(first_day, sel_sy.start_date)
            range_end = min(last_day, sel_sy.end_date)
            if range_start <= range_end:
                days = [range_start + timedelta(n) for n in range((range_end - range_start).days + 1)]

                enroll_qs = Enrollment.objects.filter(school_year=sel_sy, active=True).select_related('student', 'section')
                if not (request.user.is_staff or request.user.is_superuser):
                    enroll_qs = enroll_qs.filter(section__adviser=request.user)
                else:
                    if sel_section_id:
                        enroll_qs = enroll_qs.filter(section_id=sel_section_id)
                enrollments = list(enroll_qs)

                if enrollments:
                    recs = AttendanceSessionRecord.objects.filter(
                        enrollment__in=enrollments,
                        date__gte=range_start,
                        date__lte=range_end,
                    )
                    by_key = {(r.enrollment_id, r.date, r.session): r for r in recs}

                    rows_m, rows_f = [], []
                    mpd = [0.0 for _ in days]
                    fpd = [0.0 for _ in days]

                    for e in enrollments:
                        s = e.student
                        day_marks = []
                        day_pairs = []
                        counts = {'P': 0.0, 'A': 0.0, 'L': 0.0, 'E': 0.0}
                        for idx, d in enumerate(days):
                            am = by_key.get((e.id, d, 'AM'))
                            pm = by_key.get((e.id, d, 'PM'))
                            ams = am.status if am else ''
                            pms = pm.status if pm else ''
                            if ams:
                                counts[ams] = counts.get(ams, 0.0) + 0.5
                                if ams in PRESENT_SET and ams != 'P':
                                    counts['P'] += 0.5
                            if pms:
                                counts[pms] = counts.get(pms, 0.0) + 0.5
                                if pms in PRESENT_SET and pms != 'P':
                                    counts['P'] += 0.5
                            cell = f"{ams}/{pms}" if (ams or pms) else ''
                            day_marks.append(cell)
                            day_pairs.append({'day': d, 'mark': cell})
                            p_inc = (0.5 if (ams in PRESENT_SET) else 0.0) + (0.5 if (pms in PRESENT_SET) else 0.0)
                            if s.sex == 'M':
                                mpd[idx] += p_inc
                            else:
                                fpd[idx] += p_inc
                        row = {
                            'lrn': s.lrn or '',
                            'name': f"{s.last_name}, {s.first_name}",
                            'sex': s.sex,
                            'birthdate': s.birthdate,
                            'day_marks': day_marks,
                            'day_pairs': day_pairs,
                            'counts': counts,
                        }
                        (rows_m if s.sex == 'M' else rows_f).append(row)
                    # Compute monthly summary for preview (cached)
                    if (request.user.is_staff or request.user.is_superuser):
                        scope = f"section:{sel_section_id or 'all'}"
                    else:
                        scope = f'user:{request.user.id}'
                    cache_key = f"sf2:{sel_sy.id}:{sel_year}:{sel_month}:{scope}"
                    summary = cache.get(cache_key)
                    if summary is None:
                        summary = _compute_sf2_summary(sel_sy, sel_year, sel_month, days, enrollments, by_key)
                        cache.set(cache_key, summary, timeout=300)
                    non_school_days = list(NonSchoolDay.objects.filter(school_year=sel_sy, date__gte=range_start, date__lte=range_end).order_by('date'))
                    total_cols = len(days) + 8
                    cpd = [ (mpd[i] + fpd[i]) for i in range(len(days)) ]
            else:
                # Selected month outside the school year range â€” show message and empty preview
                messages.error(request, 'Selected month is outside the school year range.')
        except Exception:
            # On any unexpected error, keep preview empty but do not break the page
            pass

    return render(request, 'attendance/report_form.html', {
        'schoolyears': sys,
        'now': today,
        'selected_sy': sel_sy,
        'selected_sy_id': sel_sy.id if sel_sy else None,
        'selected_year': sel_year,
        'selected_month': sel_month,
        'selected_section_id': sel_section_id,
        'year_options': year_options,
        'month_options': month_options,
        'sections': list(Section.objects.filter(school_year=sel_sy)) if sel_sy else [],
        'days': days,
        'rows': rows_m + rows_f if sel_sy and enrollments else rows,
        'rows_m': locals().get('rows_m', []),
        'rows_f': locals().get('rows_f', []),
        'mpd': locals().get('mpd', []),
        'fpd': locals().get('fpd', []),
        'cpd': locals().get('cpd', []),
        'total_cols': locals().get('total_cols', (len(days) + 8 if days else 0)),
        'summary': summary,
        'non_school_days': non_school_days,
        'nsd_dates': [d.date for d in non_school_days],
    })


@login_required
def export_monthly_report(request):
    if not has_feature(request.user, 'view_reports'):
        messages.warning(request, 'You are not allowed to export reports.')
        return redirect('attendance:dashboard')
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:  # pragma: no cover
        messages.error(request, 'openpyxl is required. Please install dependencies: pip install -r requirements.txt')
        return redirect('attendance:report_form')
    except Exception as e:  # pragma: no cover
        messages.error(request, f'Error initializing Excel export: {e}')
        return redirect('attendance:report_form')

    schoolyear_id = int(request.GET.get('schoolyear_id'))
    year = int(request.GET.get('year'))
    month = int(request.GET.get('month'))
    sel_section_id = request.GET.get('section_id')
    try:
        sel_section_id = int(sel_section_id) if sel_section_id not in (None, '', 'all') else None
    except (TypeError, ValueError):
        sel_section_id = None

    sy = get_object_or_404(SchoolYear, pk=schoolyear_id)

    # Determine actual range within the selected month intersecting the school year
    first_day = date(year, month, 1)
    last_day = date(year, month, monthrange(year, month)[1])
    range_start = max(first_day, sy.start_date)
    range_end = min(last_day, sy.end_date)
    if range_start > range_end:
        messages.error(request, 'Selected month is outside the school year range.')
        return redirect('attendance:report_form')

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"SF2 {year}-{month:02d}"

    days = [d for d in (range_start + timedelta(n) for n in range((range_end - range_start).days + 1))]

    # Header row matching a simplified DepEd SF2 style
    headers = [
        'LRN', 'Learner\'s Name', 'Sex', 'Birthdate',
    ] + [str(d.day) for d in days] + ['Present', 'Absent', 'Late', 'Excused']

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    enroll_qs = Enrollment.objects.filter(school_year=sy, active=True).select_related('student', 'section')
    if not (request.user.is_staff or request.user.is_superuser):
        enroll_qs = enroll_qs.filter(section__adviser=request.user)
        if not enroll_qs.exists():
            messages.error(request, 'You have no section or students for this school year.')
            return redirect('attendance:report_form')
    else:
        if sel_section_id:
            enroll_qs = enroll_qs.filter(section_id=sel_section_id)
    enrollments = list(enroll_qs)

    # Preload all records in this month for performance
    recs = AttendanceSessionRecord.objects.filter(
        enrollment__in=enrollments,
        date__gte=range_start,
        date__lte=range_end,
    )
    by_key = {(r.enrollment_id, r.date, r.session): r for r in recs}

    # Non-school days for shading
    nsd_dates = set(NonSchoolDay.objects.filter(school_year=sy, date__gte=range_start, date__lte=range_end).values_list('date', flat=True))
    nsd_fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

    # Partition by sex for grouped output
    males = [e for e in enrollments if e.student.sex == 'M']
    females = [e for e in enrollments if e.student.sex == 'F']

    def write_group(group, label):
        nonlocal row
        mpd = [0.0 for _ in days]
        for e in group:
            s = e.student
            ws.cell(row=row, column=1, value=s.lrn or '')
            ws.cell(row=row, column=2, value=f"{s.last_name}, {s.first_name}")
            ws.cell(row=row, column=3, value=s.sex)
            ws.cell(row=row, column=4, value=s.birthdate.strftime('%Y-%m-%d') if s.birthdate else '')

            counts = {'P': 0.0, 'A': 0.0, 'L': 0.0, 'E': 0.0}
            col = 5
            for i, d in enumerate(days):
                am = by_key.get((e.id, d, 'AM'))
                pm = by_key.get((e.id, d, 'PM'))
                ams = am.status if am else ''
                pms = pm.status if pm else ''
                cell = f"{ams}/{pms}" if (ams or pms) else ''
                # Code buckets with present including L/E
                if ams:
                    counts[ams] = counts.get(ams, 0.0) + 0.5
                    if ams in PRESENT_SET and ams != 'P':
                        counts['P'] += 0.5
                if pms:
                    counts[pms] = counts.get(pms, 0.0) + 0.5
                    if pms in PRESENT_SET and pms != 'P':
                        counts['P'] += 0.5
                # Per-day present increment
                p_inc = (0.5 if (ams in PRESENT_SET) else 0.0) + (0.5 if (pms in PRESENT_SET) else 0.0)
                mpd[i] += p_inc
                c = ws.cell(row=row, column=col, value=cell)
                if d in nsd_dates:
                    c.fill = nsd_fill
                col += 1

            ws.cell(row=row, column=col, value=counts['P']); col += 1
            ws.cell(row=row, column=col, value=counts['A']); col += 1
            ws.cell(row=row, column=col, value=counts['L']); col += 1
            ws.cell(row=row, column=col, value=counts['E']); col += 1

            row += 1

        # Totals row for present per day
        ws.cell(row=row, column=2, value=f"{label} present per day").font = Font(bold=True)
        col = 5
        for v in mpd:
            ws.cell(row=row, column=col, value=v); col += 1
        row += 2
        return mpd

    row = 2
    mpd_m = write_group(males, 'Male')
    mpd_f = write_group(females, 'Female')
    # Combined row
    ws.cell(row=row, column=2, value='Combined present per day').font = Font(bold=True)
    col = 5
    for v in [ (mpd_m[i] + mpd_f[i]) for i in range(len(days)) ]:
        ws.cell(row=row, column=col, value=v); col += 1
    row += 2

    # Auto width (simple heuristic)
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max(10, min(25, length + 2))

    # Append SF2 monthly summary block
    summary = _compute_sf2_summary(sy, year, month, days, enrollments, by_key)
    row += 2
    ws.cell(row=row, column=1, value='Monthly Summary (SF2)').font = Font(bold=True)
    row += 1
    ws.cell(row=row, column=1, value='No. of School Days in month')
    ws.cell(row=row, column=2, value=summary['school_days'])
    row += 1
    ws.cell(row=row, column=2, value='M').font = Font(bold=True)
    ws.cell(row=row, column=3, value='F').font = Font(bold=True)
    ws.cell(row=row, column=4, value='TOTAL').font = Font(bold=True)
    row += 1
    def _w(label, key):
        nonlocal row
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=summary['by']['M'][key])
        ws.cell(row=row, column=3, value=summary['by']['F'][key])
        ws.cell(row=row, column=4, value=summary['by']['T'][key])
        row += 1
    _w("Enrolment as of 1st Friday", 'enrol_first_friday')
    _w("Late enrolment during the month (beyond cut-off)", 'late_enrol')
    _w("Registered learners as of end of month", 'registered_eom')
    _w("% of enrolment as of end of month", 'pct_enrol_eom')
    _w("Average Daily Attendance", 'ada')
    _w("% of attendance for the month", 'pct_attendance')
    _w("Students absent for 5 consecutive days", 'absent5')

    filename = f"SF2_{sy.name}_{year}-{month:02d}.xlsx".replace('/', '-')
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


@login_required
def report_preview(request):
    if not has_feature(request.user, 'view_reports'):
        messages.warning(request, 'You are not allowed to view reports.')
        return redirect('attendance:dashboard')
    schoolyear_id = int(request.GET.get('schoolyear_id'))
    year = int(request.GET.get('year'))
    month = int(request.GET.get('month'))
    sel_section_id = request.GET.get('section_id')
    try:
        sel_section_id = int(sel_section_id) if sel_section_id not in (None, '', 'all') else None
    except (TypeError, ValueError):
        sel_section_id = None

    sy = get_object_or_404(SchoolYear, pk=schoolyear_id)

    first_day = date(year, month, 1)
    last_day = date(year, month, monthrange(year, month)[1])
    range_start = max(first_day, sy.start_date)
    range_end = min(last_day, sy.end_date)
    if range_start > range_end:
        messages.error(request, 'Selected month is outside the school year range.')
        return redirect('attendance:report_form')

    days = [range_start + timedelta(n) for n in range((range_end - range_start).days + 1)]

    enroll_qs = Enrollment.objects.filter(school_year=sy, active=True).select_related('student', 'section')
    if not (request.user.is_staff or request.user.is_superuser):
        enroll_qs = enroll_qs.filter(section__adviser=request.user)
        if not enroll_qs.exists():
            messages.error(request, 'You have no section or students for this school year.')
            return redirect('attendance:report_form')
    else:
        if sel_section_id:
            enroll_qs = enroll_qs.filter(section_id=sel_section_id)
    enrollments = list(enroll_qs)

    # Use session-based attendance (AM/PM) like export and report_form
    recs = AttendanceSessionRecord.objects.filter(
        enrollment__in=enrollments,
        date__gte=range_start,
        date__lte=range_end,
    )
    by_key = {(r.enrollment_id, r.date, r.session): r for r in recs}

    rows = []
    rows_m, rows_f = [], []
    non_school_days = list(NonSchoolDay.objects.filter(school_year=sy, date__gte=range_start, date__lte=range_end).order_by('date'))
    mpd = [0.0 for _ in days]
    fpd = [0.0 for _ in days]
    for e in enrollments:
        s = e.student
        day_marks = []
        day_pairs = []
        counts = {'P': 0.0, 'A': 0.0, 'L': 0.0, 'E': 0.0}
        for idx, d in enumerate(days):
            am = by_key.get((e.id, d, 'AM'))
            pm = by_key.get((e.id, d, 'PM'))
            ams = am.status if am else ''
            pms = pm.status if pm else ''
            if ams:
                counts[ams] = counts.get(ams, 0.0) + 0.5
                if ams in PRESENT_SET and ams != 'P':
                    counts['P'] += 0.5
            if pms:
                counts[pms] = counts.get(pms, 0.0) + 0.5
                if pms in PRESENT_SET and pms != 'P':
                    counts['P'] += 0.5
            cell = f"{ams}/{pms}" if (ams or pms) else ''
            day_marks.append(cell)
            day_pairs.append({'day': d, 'mark': cell})
            p_inc = (0.5 if (ams in PRESENT_SET) else 0.0) + (0.5 if (pms in PRESENT_SET) else 0.0)
            if s.sex == 'M':
                mpd[idx] += p_inc
            else:
                fpd[idx] += p_inc
        row = {
            'lrn': s.lrn or '',
            'name': f"{s.last_name}, {s.first_name}",
            'sex': s.sex,
            'birthdate': s.birthdate,
            'day_marks': day_marks,
            'day_pairs': day_pairs,
            'counts': counts,
        }
        rows.append(row)
        (rows_m if s.sex == 'M' else rows_f).append(row)

    # Compute SF2 summary for preview (cached)
    if (request.user.is_staff or request.user.is_superuser):
        scope = f"section:{sel_section_id or 'all'}"
    else:
        scope = f'user:{request.user.id}'
    cache_key = f"sf2:{sy.id}:{year}:{month}:{scope}"
    summary = cache.get(cache_key)
    if summary is None:
        summary = _compute_sf2_summary(sy, year, month, days, enrollments, by_key)
        cache.set(cache_key, summary, timeout=300)

    context = {
        'schoolyear': sy,
        'year': year,
        'month': month,
        'selected_section_id': sel_section_id,
        'days': days,
        'rows': rows_m + rows_f,
        'rows_m': rows_m,
        'rows_f': rows_f,
        'mpd': mpd,
        'fpd': fpd,
        'cpd': [ (mpd[i] + fpd[i]) for i in range(len(days)) ],
        'total_cols': len(days) + 8,
        'summary': summary,
        'non_school_days': non_school_days,
        'nsd_dates': [d.date for d in non_school_days],
    }
    return render(request, 'attendance/report_preview.html', context)


@login_required
def report_day_delete(request, schoolyear_id: int, year: int, month: int, day: int):
    if not has_feature(request.user, 'view_reports'):
        messages.warning(request, 'You are not allowed to modify reports.')
        return redirect('attendance:dashboard')
    """Confirm and delete all attendance records for a specific day.
    Staff users affect the whole school year; non-staff restricted to their section(s).
    """
    sy = get_object_or_404(SchoolYear, pk=schoolyear_id)
    try:
        target_date = date(year, month, day)
    except Exception:
        messages.error(request, 'Invalid date provided.')
        return redirect('attendance:report_form')

    # Date must be inside the school year window
    if target_date < sy.start_date or target_date > sy.end_date:
        messages.error(request, 'Selected day is outside the school year range.')
        return redirect('attendance:report_form')

    # Scope of enrollments
    enroll_qs = Enrollment.objects.filter(school_year=sy, active=True)
    if not (request.user.is_staff or request.user.is_superuser):
        enroll_qs = enroll_qs.filter(section__adviser=request.user)
        if not enroll_qs.exists():
            messages.error(request, 'You have no eligible students on this school year.')
            return redirect('attendance:report_form')

    # Counts to show in confirmation
    sess_qs = AttendanceSessionRecord.objects.filter(enrollment__in=enroll_qs, date=target_date)
    per_qs = AttendancePeriodRecord.objects.filter(enrollment__in=enroll_qs, date=target_date)
    sess_count = sess_qs.count()
    per_count = per_qs.count()

    if request.method == 'POST':
        deleted_sessions = sess_count
        deleted_periods = per_count
        # Perform deletions
        per_qs.delete()
        sess_qs.delete()

        # Invalidate cached SF2 summaries for this month
        try:
            scope_keys = []
            # all-scope for staff
            scope_keys.append(f"sf2:{sy.id}:{year}:{month}:all")
            # adviser scope(s)
            adviser_ids = set(enroll_qs.values_list('section__adviser_id', flat=True))
            adviser_ids.add(getattr(request.user, 'id', None))
            for aid in adviser_ids:
                if aid:
                    scope_keys.append(f"sf2:{sy.id}:{year}:{month}:user:{aid}")
            for k in scope_keys:
                cache.delete(k)
        except Exception:
            pass

        messages.success(
            request,
            f"Deleted {deleted_sessions} session and {deleted_periods} period record(s) for {target_date}."
        )
        # Redirect back to report form with the same filters
        return redirect(f"{reverse('attendance:report_form')}?schoolyear_id={sy.id}&year={year}&month={month}")

    return render(request, 'attendance/day_confirm_delete.html', {
        'schoolyear': sy,
        'target_date': target_date,
        'sess_count': sess_count,
        'per_count': per_count,
        'year': year,
        'month': month,
    })


@login_required
def report_day_mark_nsd(request, schoolyear_id: int, year: int, month: int, day: int):
    if not has_feature(request.user, 'manage_reports'):
        messages.error(request, 'Only staff can mark Non-School Days.')
        return redirect('attendance:report_form')
    sy = get_object_or_404(SchoolYear, pk=schoolyear_id)
    try:
        target_date = date(year, month, day)
    except Exception:
        messages.error(request, 'Invalid date provided.')
        return redirect('attendance:report_form')
    if target_date < sy.start_date or target_date > sy.end_date:
        messages.error(request, 'Selected day is outside the school year range.')
        return redirect('attendance:report_form')

    existing = NonSchoolDay.objects.filter(school_year=sy, date=target_date).first()
    if request.method == 'POST':
        kind = request.POST.get('kind') or 'HOL'
        title = (request.POST.get('title') or '').strip() or ('Holiday' if kind == 'HOL' else 'Class Suspension')
        notes = (request.POST.get('notes') or '').strip()
        NonSchoolDay.objects.update_or_create(
            school_year=sy, date=target_date,
            defaults={'kind': kind, 'title': title, 'notes': notes},
        )
        # Invalidate SF2 cache for month
        try:
            scope_keys = [f"sf2:{sy.id}:{year}:{month}:all"]
            # Include all adviser scopes (best-effort)
            adviser_ids = set(Enrollment.objects.filter(school_year=sy).values_list('section__adviser_id', flat=True))
            for aid in adviser_ids:
                if aid:
                    scope_keys.append(f"sf2:{sy.id}:{year}:{month}:user:{aid}")
            for k in scope_keys:
                cache.delete(k)
        except Exception:
            pass
        messages.success(request, f'Marked {target_date} as a Non-School Day.')
        return redirect(f"{reverse('attendance:report_form')}?schoolyear_id={sy.id}&year={year}&month={month}")

    return render(request, 'attendance/nsd_mark_form.html', {
        'schoolyear': sy,
        'target_date': target_date,
        'year': year,
        'month': month,
        'existing': existing,
    })


@login_required
def report_day_unmark_nsd(request, schoolyear_id: int, year: int, month: int, day: int):
    if not has_feature(request.user, 'manage_reports'):
        messages.error(request, 'Only staff can unmark Non-School Days.')
        return redirect('attendance:report_form')
    sy = get_object_or_404(SchoolYear, pk=schoolyear_id)
    try:
        target_date = date(year, month, day)
    except Exception:
        messages.error(request, 'Invalid date provided.')
        return redirect('attendance:report_form')
    if target_date < sy.start_date or target_date > sy.end_date:
        messages.error(request, 'Selected day is outside the school year range.')
        return redirect('attendance:report_form')

    obj = NonSchoolDay.objects.filter(school_year=sy, date=target_date).first()
    if not obj:
        messages.info(request, 'This day is not marked as a Non-School Day.')
        return redirect(f"{reverse('attendance:report_form')}?schoolyear_id={sy.id}&year={year}&month={month}")

    if request.method == 'POST':
        obj.delete()
        # Invalidate SF2 cache for month
        try:
            scope_keys = [f"sf2:{sy.id}:{year}:{month}:all"]
            adviser_ids = set(Enrollment.objects.filter(school_year=sy).values_list('section__adviser_id', flat=True))
            for aid in adviser_ids:
                if aid:
                    scope_keys.append(f"sf2:{sy.id}:{year}:{month}:user:{aid}")
            for k in scope_keys:
                cache.delete(k)
        except Exception:
            pass
        messages.success(request, f'Unmarked {target_date} as a Non-School Day.')
        return redirect(f"{reverse('attendance:report_form')}?schoolyear_id={sy.id}&year={year}&month={month}")

    return render(request, 'attendance/nsd_unmark_confirm.html', {
        'schoolyear': sy,
        'target_date': target_date,
        'year': year,
        'month': month,
        'obj': obj,
    })

@login_required
def non_school_days_import(request):
    if not has_feature(request.user, 'manage_reports'):
        messages.error(request, 'You are not allowed to import Non-School Days.')
        return redirect('attendance:report_form')

    sys = SchoolYear.objects.all()
    if request.method == 'POST':
        try:
            schoolyear_id = int(request.POST.get('schoolyear_id'))
        except (TypeError, ValueError):
            schoolyear_id = None
        sy = get_object_or_404(SchoolYear, pk=schoolyear_id) if schoolyear_id else None
        file = request.FILES.get('file')
        if not sy or not file:
            messages.error(request, 'Please select a School Year and choose a CSV file to upload.')
            return redirect('attendance:non_school_days_import')
        try:
            wrapper = TextIOWrapper(file.file, encoding='utf-8-sig')
            reader = csv.DictReader(wrapper)
        except Exception:
            messages.error(request, 'Invalid CSV file. Ensure it has a header: date,kind,title,notes')
            return redirect('attendance:non_school_days_import')

        kind_map = {
            'hol': 'HOL', 'holiday': 'HOL', 'h': 'HOL',
            'sus': 'SUS', 'suspension': 'SUS', 'class suspension': 'SUS', 'c': 'SUS'
        }

        created = 0
        updated = 0
        skipped = 0
        for row in reader:
            raw_date = (row.get('date') or '').strip()
            raw_kind = (row.get('kind') or '').strip().lower()
            title = (row.get('title') or '').strip()
            notes = (row.get('notes') or '').strip()
            if not raw_date or not title:
                skipped += 1
                continue
            try:
                y, m, d = [int(x) for x in raw_date.split('-')]
                dt = date(y, m, d)
            except Exception:
                skipped += 1
                continue
            if dt < sy.start_date or dt > sy.end_date:
                skipped += 1
                continue
            kind = kind_map.get(raw_kind, 'HOL')
            _, was_created = NonSchoolDay.objects.update_or_create(
                school_year=sy, date=dt,
                defaults={'kind': kind, 'title': title, 'notes': notes},
            )
            if was_created:
                created += 1
            else:
                updated += 1
        messages.success(request, f'Imported: created {created}, updated {updated}, skipped {skipped}.')
        return redirect('attendance:report_form')

    return render(request, 'attendance/non_school_days_import.html', {
        'schoolyears': sys,
    })

